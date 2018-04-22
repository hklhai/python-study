# coding=utf-8
import openpyxl

def cpoy_sheet():
    """
    python3读写excel
    """
    # 读取数据
    wb1 = openpyxl.load_workbook(r'C:\Users\lenovo\Desktop\excel\A.xlsx')
    wb2 = openpyxl.load_workbook(r'C:\Users\lenovo\Desktop\excel\ACOPY.xlsx')
    sheets1 = wb1.sheetnames  # 获取sheet页
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]

    max_row = sheet1.max_row  # 最大行数
    max_column = sheet1.max_column  # 最大列数

    for m in range(1, max_row + 1):
        for n in range(97, 97 + max_column):  # chr(97)='a'
            n = chr(n)  # ASCII字符
            i = '%s%d' % (n, m)  # 单元格编号
            cell1 = sheet1[i].value  # 获取data单元格数据
            sheet2[i].value = cell1  # 赋值到test单元格

    wb2.save(r'C:\Users\lenovo\Desktop\excel\ACOPY.xlsx')  # 保存数据
    wb1.close()  # 关闭excel
    wb2.close()


if __name__ == "__main__":
    cpoy_sheet()
